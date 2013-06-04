#!/usr/bin/env python
#
# BitCurator
# Copyright (C) 2012
# All rights reserved.
# 
# This code is distributed under the terms of the GNU General Public 
# License, Version 3. See the text file "COPYING" for further details 
# about the terms of this license.
#
# Generate report in Excel format (from xml input)
# 

import sys,os,shelve
import re,dfxml,fiwalk
from bc_utils import filename_from_path
from bc_utils import is_special_file

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter

def bc_generate_feature_xlsx(PdfReport, data, feature_file):

    wb = Workbook(optimized_write = True)
    dest_filename = PdfReport.featuredir +'/'+ (filename_from_path(feature_file))[10:-3] + "xlsx"
    row_idx = [2]
    ##ws = wb.worksheets[0]
    ws = wb.create_sheet()
    ws.title = "File Feature Information"

    ws.cell('%s%s'%('A', '1')).value = '%s' % "Filename"
    ws.cell('%s%s'%('B', '1')).value = '%s' % "Feature"
    ws.cell('%s%s'%('C', '1')).value = '%s' % "Position"

    linenum=0
    for row in data:
        # Skip the lines with known text lines to be eliminated
        if (re.match("Total features",str(row))):
           continue
        filename = "Unknown"
        feature = "Unknown"
        position = "Unknown"

        # Some lines in the annotated_xxx.txt have less than three
        # columns where filename or feature may be missing.
        if len(row) > 3:
                filename = row[3]
        else:
                filename = "Unknown"

        if len(row) > 1:
                feature = row[1]
        else:
                feature = "Unknown"

        position = row[0]

        # If it is a special file, check if the user wants it to
        # be repoted. If not, exclude this from the table.
        if (PdfReport.bc_config_report_special_files == False) and \
                            (is_special_file(filename)):
                ## print("D: File %s is special. So skipping" %(filename))
                continue
        ws.cell('%s%s'%('A', row_idx[0])).value = '%s' % filename
        ws.cell('%s%s'%('B', row_idx[0])).value = '%s' % feature
        ws.cell('%s%s'%('C', row_idx[0])).value = '%s' % position

        row_idx[0] += 1

    wb.save(filename=dest_filename)


